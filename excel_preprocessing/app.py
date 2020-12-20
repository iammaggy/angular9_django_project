import os
import string
import pandas as pd
import numpy as np


def excel_processing():
    # Read data from excelsheet
    df = pd.read_excel("/home/mangeshkadam/Desktop/practice/BOM file for Data processing.xlsx", sheet_name="Source",
                       engine="openpyxl")
    df_final = df.replace('nan', np.nan)
    df_final = df.replace({np.nan: None})
    # Get unique Item name
    df_unique_item = df_final['Item Name'].unique()
    # Traverse through loop and remove None value
    for item in df_unique_item[df_unique_item != None]:
        df_details = df_final[df_final['Item Name'] == item]
        # Get unique Level
        df_unique_level = df_details['Level'].unique()
        for data in df_unique_level[df_unique_level != None]:
            # Filter data as per the selected level and create sheet
            if data == '.1':
                df_data = df_details[df_details['Level'] == data]
                design_excelsheet(df_data.to_dict('records'), df_data['Item Name'].iloc[0])
            else:
                df_data = df_details[df_details['Level'] == data]
                # Get the parent Item Name of the child
                df_data['Item Name'] = df_details.loc[df_details['Level'].shift(-1) == data, 'Raw material'].iloc[0]
                design_excelsheet(df_data.to_dict('records'), df_data['Item Name'].iloc[0])


def design_excelsheet(data_list, sheetname):
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.styles.borders import Border, Side
    from openpyxl import Workbook
    from openpyxl import load_workbook

    thick_border_all = Border(left=Side(style='thick', color='000000'),
                              top=Side(style='thick', color='000000'),
                              bottom=Side(style='thick', color='000000'),
                              right=Side(style='thick', color='000000'),
                              )
    border_all = Border(left=Side(style='thin', color='2E2827'),
                        top=Side(style='thin', color='2E2827'),
                        bottom=Side(style='thin', color='2E2827'),
                        right=Side(style='thin', color='2E2827'),
                        )

    left_bottom_thick = Border(left=Side(style='thick', color='000000'),
                               top=Side(style='thin', color='2E2827'),
                               bottom=Side(style='thick', color='000000'),
                               right=Side(style='thin', color='2E2827'),
                               )

    right_bottom_thick = Border(left=Side(style='thin', color='2E2827'),
                                top=Side(style='thin', color='2E2827'),
                                bottom=Side(style='thick', color='000000'),
                                right=Side(style='thick', color='000000'),
                                )

    left_thick = Border(left=Side(style='thick', color='000000'),
                        top=Side(style='thin', color='2E2827'),
                        bottom=Side(style='thin', color='2E2827'),
                        right=Side(style='thin', color='2E2827')
                        )

    right_thick = Border(left=Side(style='thin', color='2E2827'),
                         top=Side(style='thin', color='2E2827'),
                         bottom=Side(style='thin', color='2E2827'),
                         right=Side(style='thick', color='000000')
                         )

    bottom_thick = Border(left=Side(style='thin', color='2E2827'),
                          top=Side(style='thin', color='2E2827'),
                          bottom=Side(style='thick', color='000000'),
                          right=Side(style='thin', color='2E2827')
                          )
    font = Font(size=12, bold=True)
    # ws = wb.active
    path = os.path.join('/home/mangeshkadam/Desktop/practice', 'bom_process_file.xlsx')
    if os.path.exists(path):
        # load workbook if path exist
        wb = load_workbook(path)
        # Create new sheet
        ws = wb.create_sheet(sheetname, 0)
    else:
        # Create workbook if path is not exist
        wb = Workbook()
        # Create new sheet
        ws = wb.create_sheet(sheetname, 0)
    ws['A1'] = 'Finished Good List'
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A5'] = 'Raw Material List'
    ws['A5'].font = Font(size=14, bold=True)
    ws.merge_cells('A5:D5')
    header_list = ['#', 'Item Description', 'Quantity', 'Unit']
    letters = [x for x in string.ascii_uppercase]
    temp_alpha = [x + y for x in string.ascii_uppercase for y in string.ascii_uppercase]
    letters.extend(temp_alpha)

    i = 0
    # Write header of the column
    for item in range(0, len(header_list)):
        ws[letters[i] + str(2)] = header_list[item]
        ws[letters[i] + str(2)].font = font
        ws[letters[i] + str(2)].fill = PatternFill(fgColor="a1bdd5", fill_type="solid")
        ws[letters[i] + str(2)].border = thick_border_all
        ws[letters[i] + str(2)].alignment = Alignment(horizontal='center')
        ws[letters[i] + str(2)].alignment = Alignment(horizontal='center')
        i = i + 1

    ws['A' + str(3)] = 1
    ws['A' + str(3)].border = left_bottom_thick
    ws['A' + str(3)].alignment = Alignment(horizontal='center')
    ws['A' + str(3)].alignment = Alignment(horizontal='center')

    ws['B' + str(3)] = sheetname
    ws['B' + str(3)].border = bottom_thick
    ws['B' + str(3)].fill = PatternFill(fgColor="fcea36", fill_type="solid")
    ws['B' + str(3)].alignment = Alignment(horizontal='center')
    ws['B' + str(3)].alignment = Alignment(horizontal='center')

    ws['C' + str(3)] = 1
    ws['C' + str(3)].border = bottom_thick
    ws['C' + str(3)].alignment = Alignment(horizontal='center')
    ws['C' + str(3)].alignment = Alignment(horizontal='center')

    ws['D' + str(3)] = 'Pc'
    ws['D' + str(3)].border = right_bottom_thick
    ws['D' + str(3)].alignment = Alignment(horizontal='center')
    ws['D' + str(3)].alignment = Alignment(horizontal='center')

    i = 0
    for item in range(0, len(header_list)):
        ws[letters[i] + str(6)] = header_list[item]
        ws[letters[i] + str(6)].font = font
        ws[letters[i] + str(6)].fill = PatternFill(fgColor="a1bdd5", fill_type="solid")
        ws[letters[i] + str(6)].border = thick_border_all
        ws[letters[i] + str(6)].alignment = Alignment(horizontal='center')
        ws[letters[i] + str(6)].alignment = Alignment(horizontal='center')
        i = i + 1

    j = 7
    index_val = 1
    # Write the data in excelsheet as per the filter data
    for data in data_list:
        col_list = ['Index', 'Raw material', 'Quantity', 'Unit']
        i = 0
        for item in col_list:
            if item == 'Index':
                ws[letters[i] + str(j)] = index_val
                ws[letters[i] + str(j)].border = left_bottom_thick if index_val == len(data_list) else left_thick
                ws[letters[i] + str(j)].alignment = Alignment(horizontal='center')
            else:
                ws[letters[i] + str(j)] = data[item]
                if item == 'Unit':
                    ws[letters[i] + str(j)].border = right_bottom_thick if index_val == len(data_list) else right_thick
                else:
                    ws[letters[i] + str(j)].border = bottom_thick if index_val == len(data_list) else border_all
                ws[letters[i] + str(j)].alignment = Alignment(horizontal='center')
                ws[letters[i] + str(j)].fill = PatternFill(fgColor="fcea36", fill_type="solid")
            i = i + 1
        index_val = index_val + 1
        j = j + 1
    wb.save(path)


excel_processing()
